import sys
import openpyxl
from rich.console import Console
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from pprint import pprint
import os
import shutil
from datetime import datetime

def dier(msg):
    pprint(msg)
    sys.exit(10)

console = Console()

gegenstaende_und_preise = {
    "Besprechungsstuhl Fiore Vierbeiner weiß/hellgrün mit Klapptisch": { "price": 241.45, "serial_number_required": False },
    "Rollcontainer 9 HE 1-2-3-3 Buche": {"price": 185.42, "serial_number_required": False },
    "Sitz-Steharbeitsplatz 180x80x64-125-Buche": { "price": 487.55, "serial_number_required": False },
    "Drehstuhl AJ5786 schwarz": { "price": 322.24, "serial_number_required": False },
    "Besprechungsstühle Cay Vierbeiner grau/hellgrün": { "price": 168.45, "serial_number_required": False },
    "Lenovo ThinkPad T14 AMD Gen 3": { "price": 2152.71, "serial_number_required": True },
    "Monitor Lenovo ThinkVision T27h-2L": { "price": 304.00, "serial_number_required": True }
}

# TODO: Barcodes für Thinkpads mit einlesen
# Bugs:
# Fügt die an der falschen Stelle ein und dann grün
# Man kann den Namen nicht ändern

PREDEFINED_ITEM_TYPES = list(gegenstaende_und_preise.keys())

gebaeude_id = "3331"
kostenstelle = "2340200G"
waehrung = "EUR"
current_person = ""
current_room = ""

def ask_for_anlagenbezeichnung():
    console.print("[yellow]Gerätetyp auswählen:[/yellow]")
    for idx, name in enumerate(PREDEFINED_ITEM_TYPES, start=1):
        console.print(f"[underline]{idx}[/underline]: {name}")
    choice = input("Nummer eingeben: ").strip()
    try:
        index = int(choice) - 1
        if 0 <= index < len(PREDEFINED_ITEM_TYPES):
            return PREDEFINED_ITEM_TYPES[index]
        else:
            console.print("[red]Ungültige Auswahl![/red]")
            return ask_for_anlagenbezeichnung()
    except ValueError:
        console.print("[red]Bitte eine gültige Zahl eingeben![/red]")
        return ask_for_anlagenbezeichnung()

def find_entry(sheet, anlagennummer):
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        if str(row[0].value).strip() == anlagennummer:
            console.print(f"[green]Gefunden: Anlagennummer {anlagennummer} in Zeile {row[0].row}[/green]")
            return row
    console.print(f"[red]Nicht gefunden: Anlagennummer {anlagennummer}[/red]")
    return None

def insert_sorted_row(sheet, anlagennummer, anlagenbezeichnung, preis):
    yellow_fill = PatternFill(
        fill_type="solid",
        start_color="FFFF00",  # RGB für Gelb
        end_color="FFFF00"
    )

    serial_number = None

    if gegenstaende_und_preise[anlagenbezeichnung]["serial_number_required"]:
        serial_number = input("Seriennummer: ")

    preis = gegenstaende_und_preise[anlagenbezeichnung]["price"]

    # Spalten: A = Inventarnummer, E = Bezeichnung, H = Währung, I = Standort, J = Raum, L = Inventurhinweis, M = Kostenstelle
    new_row = [anlagennummer, None, None, None, anlagenbezeichnung, serial_number, preis, waehrung, gebaeude_id, current_room, None, current_person, kostenstelle]
    inserted = False

    for row_idx in range(2, sheet.max_row + 1):
        cell_anlagennummer = sheet.cell(row=row_idx, column=1).value
        if cell_anlagennummer is None or str(cell_anlagennummer).strip() == "":
            continue

        if int(anlagennummer) < int(cell_anlagennummer):
            sheet.insert_rows(row_idx)
            for col, val in enumerate(new_row, start=1):
                sheet.cell(row=row_idx, column=col, value=val)

                if col == 7:
                    cell.number_format = '#,##0.00'
            inserted = True
            console.print(f"[blue]Neue Zeile sortiert eingefügt vor Zeile {row_idx}[/blue]")

            # Gelbes Füllmuster anwenden
            for col in range(1, len(new_row) + 1):
                cell = sheet.cell(row=row_idx, column=col).fill = yellow_fill

                if col == 7:
                    cell.number_format = '#,##0.00'

            break

    if not inserted:
        # Finde die letzte Zeile, in der Spalte A (column=1) ein Wert steht
        last_data_row = 0
        for row_idx in range(2, sheet.max_row + 1):
            cell_anlagennummer = sheet.cell(row=row_idx, column=1).value
            if cell_anlagennummer is not None and str(cell_anlagennummer).strip() != "":
                last_data_row = row_idx

        # Neue Zielzeile ist eine Zeile darunter
        insert_row_index = last_data_row + 1

        # Neue Zeile einfügen
        sheet.insert_rows(insert_row_index)
        for col_index, value in enumerate(new_row, start=1):
            sheet.cell(row=insert_row_index, column=col_index, value=value)

        # Anschaffungswert in Spalte C (Spalte G) setzen
        try:
            cell = sheet.cell(row=insert_row_index, column=7, value=preis)
            cell.number_format = '#,##0.00'
        except Exception as e:
            console.print(f"[red]Fehler beim Setzen des Preises: {e}[/red]")

        console.print(f"[blue]Neue Zeile bei Zeile {insert_row_index} eingefügt[/blue]")

        # Gelbes Füllmuster anwenden
        for col in range(1, len(new_row) + 1):
            try:
                sheet.cell(row=insert_row_index, column=col).fill = yellow_fill
            except Exception as e:
                console.print(f"[red]Fehler beim Anwenden des gelben Füllmusters in Spalte {col}: {e}[/red]")

    console.print(f"[blue]Eintrag: {anlagennummer}, {anlagenbezeichnung}, Wert: {preis}, Währung: {waehrung}, "
                  f"Standort: {gebaeude_id}, Raum: {current_room}, Person: {current_person}[/blue]")

def get_unique_filename(path):
    """
    Wenn die Datei 'path' existiert, wird ein neuer Name mit -1, -2 usw. generiert.
    """
    directory, original_filename = os.path.split(path)
    name, ext = os.path.splitext(original_filename)
    counter = 1
    new_path = path

    while os.path.exists(new_path):
        new_filename = f"{name}-{counter}{ext}"
        new_path = os.path.join(directory, new_filename)
        counter += 1

    return new_path

def save_workbook(wb, file_name):
    try:
        # Backup-Verzeichnis erstellen
        backup_dir = os.path.join(os.getcwd(), "python_script_backups", datetime.today().strftime('%Y-%m-%d'))
        if not os.path.exists(backup_dir):
            os.makedirs(backup_dir)

        # Backup der bestehenden Datei (wenn sie existiert)
        if os.path.isfile(file_name):
            base_name = os.path.basename(file_name)
            backup_path = os.path.join(backup_dir, base_name)
            unique_backup_path = get_unique_filename(backup_path)

            try:
                shutil.copy2(file_name, unique_backup_path)
                console.print(f"[yellow]Backup gespeichert unter: {unique_backup_path}[/yellow]")
            except Exception as e:
                console.print(f"[red]Fehler beim Erstellen des Backups: {e}[/red]")

        # Arbeitsmappe speichern
        wb.save(file_name)
        console.print(f"[green]Änderungen erfolgreich gespeichert in {file_name}[/green]")

    except Exception as e:
        console.print(f"[red]Fehler beim Speichern der Datei: {e}[/red]")

def mark_row_as_confirmed(sheet, row_idx):
    # Grüne Farbe im hex Format (Grün ohne Alpha)
    green_fill = PatternFill(
        fill_type="solid",
        start_color="00FF00",  # Hex für Grün
        end_color="00FF00"     # Hex für Grün
    )

    # Zelle in Spalte A und der Zeile row_idx auswählen
    cell = sheet.cell(row=row_idx, column=1)

    # Anwenden der grünen Füllung
    cell.fill = green_fill

    console.print(f"[green]Zelle A{row_idx} erfolgreich grün überschrieben.[/green]")

def main():
    global current_person, current_room

    if len(sys.argv) != 3:
        console.print("[red]Verwendung: python excel.py <excelfile> <worksheet>[/red]")
        sys.exit(1)

    excel_file = sys.argv[1]
    worksheet_name = sys.argv[2]

    try:
        wb = openpyxl.load_workbook(excel_file)
        sheet = wb[worksheet_name]
        console.print(f"[green]Arbeitsblatt '{worksheet_name}' wurde geöffnet.[/green]")
    except Exception as e:
        console.print(f"[red]Fehler beim Öffnen der Excel-Datei: {e}[/red]")
        sys.exit(1)

    while not current_room:
        current_room = input("Bitte geben Sie die Raumnummer ein: ").strip()
        if not current_room:
            console.print("[red]Raumnummer darf nicht leer sein.[/red]")

    while not current_person:
        current_person = input("Bitte Namen der zugeordneten Person eingeben: ").strip()
        if not current_person:
            console.print("[red]Name darf nicht leer sein.[/red]")

    while True:
        ask_for_input = "Bitte geben Sie die Anlagennummer ein (oder 'q'=Beenden, 'p'=Person ändern, 'r'=Raum ändern): "
        if current_person:
            ask_for_input = f"Name: {current_person}, {ask_for_input}"
        if current_room:
            ask_for_input = f"Raum: {current_room}, {ask_for_input}"

        anlagennummer_oder_kommando = input(ask_for_input).strip()

        if anlagennummer_oder_kommando.lower() == 'q':
            console.print("[yellow]Beenden...[/yellow]")
            break
        if anlagennummer_oder_kommando.lower() == 'p':
            current_person = input("Bitte neuen Namen der Person eingeben: ").strip()
            if current_person:
                console.print(f"[green]Person geändert zu: {current_person}[/green]")
            else:
                console.print("[red]Name darf nicht leer sein.[/red]")
            continue
        if anlagennummer_oder_kommando.lower() == 'r':
            current_room = input("Neue Raumnummer eingeben: ").strip()
            if current_room:
                console.print(f"[green]Raumnummer geändert zu: {current_room}[/green]")
            else:
                console.print("[red]Raumnummer darf nicht leer sein.[/red]")
            continue

        row = find_entry(sheet, anlagennummer_oder_kommando)

        if row:
            headers = []

            for cell in sheet[1]:
                headers.append(cell.value)

            console.print("Zeile enthält die folgenden Daten:")

            for i in range(len(headers)):
                header = f"Spalte {i + 1}"
                if i < len(headers):
                    header = headers[i]

                value = ""
                if i < len(row):
                    value = row[i].value

                console.print(f"[cyan]{header}[/cyan]: {value}")

            edit_msg = "Ist das korrekt? (Enter für Ja, 'e' zum Bearbeiten): "

            action = input(edit_msg).strip()

            is_valid_option = False

            while not is_valid_option:
                if action.lower() == "e":
                    console.print("[yellow]Welche Option möchtest du bearbeiten?[/yellow]")
                    print("p: Person")
                    print("r: Raum")
                    print("s: Seriennummer")
                    print("z: Zurück")
                    option = input("Gebe die Nummer der zu bearbeitenden Option ein: ").strip()

                    if option.lower() == "p":
                        row[11].value = current_person
                        console.print(f"[blue]Person geändert auf: {current_person}[/blue]")
                        mark_row_as_confirmed(sheet, row[0].row)
                        save_workbook(wb, excel_file)
                        is_valid_option = True

                    elif option.lower() == "s":
                        serial_number = input("Seriennummer: ")

                        sheet.cell(row=row[0].row, column=6, value=serial_number)
                        console.print(f"[blue]Seriennummer (Spalte F) geändert auf: {serial_number}[/blue]")
                        mark_row_as_confirmed(sheet, row[0].row)
                        save_workbook(wb, excel_file)
                        is_valid_option = True

                    elif option.lower() == "r":
                        sheet.cell(row=row[0].row, column=10, value=current_room)
                        console.print(f"[blue]Raum (Spalte J) geändert auf: {current_room}[/blue]")
                        mark_row_as_confirmed(sheet, row[0].row)
                        save_workbook(wb, excel_file)
                        is_valid_option = True

                    elif option.lower() == "z":
                        console.print("[green]Ich änder doch nix[/green]")
                        is_valid_option = True

                    else:
                        console.print(f"[red]Ungültige Option! Option: {option}[/red]")

                elif action.lower() in ["", "y", "j"]:
                    console.print("[green]Bestätigt. Keine Änderungen.[/green]")
                    mark_row_as_confirmed(sheet, row[0].row)
                    save_workbook(wb, excel_file)
                    is_valid_option = True

                else:
                    console.print(f"[red]Ungültige Eingabe! Eingabe: '{action}'[/red]")
                    action = input(edit_msg).strip()
                    is_valid_option = False


        else:
            anlagenbezeichnung = ask_for_anlagenbezeichnung()

            preis = gegenstaende_und_preise[anlagenbezeichnung]["price"]

            insert_sorted_row(sheet, anlagennummer_oder_kommando, anlagenbezeichnung, preis)
            save_workbook(wb, excel_file)

if __name__ == "__main__":
    try:
        main()
    except (EOFError, KeyboardInterrupt):
        console.print("\n[yellow]Du hast das Programm beendet[/yellow]")
